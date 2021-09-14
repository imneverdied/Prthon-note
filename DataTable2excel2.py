import openpyxl
# pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
letter = 'ABCDEFGHIJKLMNOPQRSTUVXYZ'
excel = Workbook()
# sheet = excel.active  # 建立一個工作中表
sheet = excel.create_sheet('ETF sheet', 0)  # 創表'ETF sheet'，放在第一頁
# 可參考資料 https://openpyxl.readthedocs.io/en/stable/tutorial.html
std = excel.get_sheet_by_name('Sheet')  # 刪掉預設Sheet
excel.remove_sheet(std)

ws = excel[excel.sheetnames[0]]
# 調整列寬
ws.column_dimensions['A'].width = 30.0
ws.column_dimensions['D'].width = 15.0

ws['A1'].fill = PatternFill("solid", fgColor="C5D9F1")
ws['B1'].fill = PatternFill("solid", fgColor="6699FF")
ws['C1'].fill = PatternFill("solid", fgColor="C5D9F1")
ws['D1'].fill = PatternFill("solid", fgColor="6699FF")


class DataTable2excel():
    def createTable(columns):
        # 欄位名稱定義
        ALL_C = columns.split(',')
        i = 0
        AAA = len(ALL_C)
        for i in range(AAA):
            index = letter[i % 26]  # 欄位數上限暫時26個
            index = index + '1'
            sheet[index] = ALL_C[i]
            i = i + 1

    def insert2Table(dataA, dataB, dataC, dataD):
        資料A = dataA
        資料B = dataB
        資料C = dataC
        資料D = dataD
        sheet.append([資料A, 資料B, 資料C, 資料D])

    def OutExcel(ExcelName):
        try:
            excel.save(ExcelName + '.xlsx')
            print('產出 ' + ExcelName + '.xlsx')
        except:
            print('excel產出錯誤')
